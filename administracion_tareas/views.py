import os
from django.shortcuts import render, redirect
from django.contrib import messages
from .forms import SesionForm, AsistenteForm, ArchivoForm
from .models import Sesion, Asistente,Reunion
from datetime import datetime   

from odf.opendocument import load
from docx import Document
from openpyxl import load_workbook
import PyPDF2
from django.core.files.storage import default_storage
from django.conf import settings
from pdfminer.high_level import extract_text
from io import BytesIO
from reportlab.pdfgen import canvas
from django.http import HttpResponse
from django.conf import settings
from pdfrw import PdfReader, PdfDict, PdfWriter

from django.core.files.uploadedfile import InMemoryUploadedFile
import tempfile

# Librería para expresiones regulares
import re
from django.contrib import messages

def agregar_sesion(request):
    if request.method == 'POST':
        form = SesionForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(
                request, 'La sesión ha sido agregada exitosamente.')
            return redirect('agregar_sesion')
    else:
        form = SesionForm()

    return render(request, 'agregar_sesion.html', {'form': form})


def agregar_asistente(request):
    if request.method == 'POST':
        form = AsistenteForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(
                request, 'El asistente ha sido agregado exitosamente.')
            return redirect('agregar_asistente')
    else:
        form = AsistenteForm()

    return render(request, 'agregar_asistente.html', {'form': form})


def lista_sesiones(request):
    sesiones = Sesion.objects.all()
    return render(request, 'lista_sesiones.html', {'sesiones': sesiones})


def lista_asistentes(request):
    asistentes = Asistente.objects.all()
    return render(request, 'lista_asistentes.html', {'asistentes': asistentes})


def procesar_archivo_odt(archivo):
    doc = load(archivo)
    contenido = ""
    for paragraph in doc.getElementsByType('text:p'):
        contenido += paragraph.text + "\n"
    return contenido


def procesar_archivo_word(archivo):
    doc = Document(archivo)
    contenido = ""
    for paragraph in doc.paragraphs:
        contenido += paragraph.text + "\n"
    return contenido


def procesar_archivo_excel(archivo):
    wb = load_workbook(archivo)
    contenido = ""
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows(values_only=True):
            contenido += "\t".join(str(cell) for cell in row) + "\n"
    return contenido


# Función para extraer y filtrar datos de un PDF
def procesar_archivo_pdf(archivo):
    # Guardamos el archivo en un directorio temporal
    with tempfile.NamedTemporaryFile(delete=False) as tmpfile:
        for chunk in archivo.chunks():
            tmpfile.write(chunk)

    # se obtiene la ruta del archivo temporal
    archivo_temporal = tmpfile.name

    # Ahora, se usa el archivo_temporal con pdfminer.six
    texto = extract_text(archivo_temporal)
    
    # Expresiones regulares para buscar los datos específicos
    patron_numero_sesion = re.compile(r'Nº sesión: \[(.*?)\]')
    patron_fecha = re.compile(r'Fecha: \[(.*?)\]')
    patron_hora_inicio = re.compile(r'Hora de inicio: \[(.*?)\]')
    patron_hora_finalizacion = re.compile(r'Hora de finalización: \[(.*?)\]|Hora de finalización: (.*?)\n')
    patron_lugar = re.compile(r'Lugar: \[(.*?)\]')
    patron_temas = re.compile(r'TEMAS TRATADOS según el orden del día establecido:\n(.*?)\n\nACUERDOS ADOPTADOS', re.DOTALL)
    patron_acuerdos = re.compile(r'ACUERDOS ADOPTADOS:(.*?)No habiendo más asuntos que tratar, se da por concluida la sesión.', re.DOTALL)

    numero_sesion = re.search(patron_numero_sesion, texto)
    fecha = re.search(patron_fecha, texto)
    hora_inicio = re.search(patron_hora_inicio, texto)
    hora_finalizacion = re.search(patron_hora_finalizacion, texto)
    lugar = re.search(patron_lugar, texto)
    temas_tratados = re.search(patron_temas, texto)
    acuerdos = re.search(patron_acuerdos, texto)

    resultado = {
        'numero_sesion': numero_sesion.group(1).strip() if numero_sesion else None,
        'fecha': fecha.group(1).strip() if fecha else None,
        'hora_inicio': hora_inicio.group(1).strip() if hora_inicio else None,
        'hora_finalizacion': hora_finalizacion.group(1).strip() if hora_finalizacion else None,
        'lugar': lugar.group(1).strip() if lugar else None,
        'temas_tratados': temas_tratados.group(1).strip() if temas_tratados else None,
        'acuerdos': acuerdos.group(1).strip() if acuerdos else None
    }
    # Después de procesar el archivo, eliminamos el archivo temporal
    os.remove(archivo_temporal)
    return resultado

def cargar_archivo(request):
    if request.method == 'POST':
        form = ArchivoForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo']  # Obtén el archivo del formulario

            # Llamamos a la función para procesar el archivo PDF
            contenido = procesar_archivo_pdf(archivo)

            if contenido:
                reunion = Reunion(
                    numero_sesion=contenido['numero_sesion'],
                    fecha=datetime.strptime(contenido['fecha'], '%d/%m/%Y').strftime('%Y-%m-%d'),
                    hora_inicio=contenido['hora_inicio'],
                    hora_finalizacion=contenido['hora_finalizacion'],
                    lugar=contenido['lugar'],
                    temas_tratados=contenido['temas_tratados'],
                    acuerdos_adoptados=contenido['acuerdos']
                )
                reunion.save() #guardar datos reunión
                # Mostrar resultado en una vista
                return render(request, 'mostrar_contenido.html', {'contenido': contenido})
            else:
                return render(request, 'error.html', {'mensaje': 'No se pudieron extraer datos del PDF.'})
    else:
        form = ArchivoForm()

    return render(request, 'cargar_archivo.html', {'form': form})

def replace_keywords_with_data(pdf_template_path, replacements):
    template = PdfReader(pdf_template_path)
    for page in template.pages:
        annotations = page['/Annots'] or []
        for annotation in annotations:
            key = annotation['/T'][1:-1]  # Nombre del campo
            if key in replacements:
                annotation.update(PdfDict(V='{}'.format(replacements[key])))

    output = BytesIO()
    PdfWriter().write(output, template)
    output.seek(0)
    return output

def create_pdf(request):
    template_path = os.path.join(settings.MEDIA_ROOT, 'plantilla.pdf')

    if request.method == 'POST':
        form = SesionForm(request.POST)
        if form.is_valid():
            sesion = form.save()

            replacements = {
                'fecha_completa': str(sesion.fecha),
                'lugar': sesion.lugar,
                'temas_tratados': sesion.temas_tratados,
                'acuerdos_adoptados': sesion.acuerdos_adoptados,
                'hora_inicio': str(sesion.hora_inicio),
                'hora_finalizacion': str(sesion.hora_finalizacion),
                'lugar_sesion': sesion.lugar,
                'cargo_autoridad': sesion.presidencia,
                'convocatoria': sesion.convocatoria,
                'organo_reunido': sesion.organo_reunido,
                'tipo_sesion': sesion.tipo_sesion,
                'numero_sesion': str(sesion.numero_sesion),
                'asistentes': sesion.asistentes
            }

            pdf_in_memory = replace_keywords_with_data(template_path, replacements)

            response = HttpResponse(pdf_in_memory.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="sesion_{sesion.numero_sesion}.pdf"'
            return response

    else:
        form = SesionForm()

    return render(request, 'form_template.html', {'form': form})